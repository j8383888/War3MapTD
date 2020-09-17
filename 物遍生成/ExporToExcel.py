# -*-coding:utf-8-*-
from openpyxl import Workbook
import openpyxl,os
import time


def writeDoo(file1, wb):
    # t = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime())
    # print(t)
    # wb = Workbook()


    if file1 == "unitOrigin.ini":
        sheet = wb["单位"]
    elif file1 == "abilityOrigin.ini":
        sheet = wb["技能"]
    elif file1 == "itemOrigin.ini":
        sheet = wb["物品"]
    elif file1 == "upgradeOrigin.ini":
        sheet = wb["科技"]

    sheet.delete_rows(1,sheet.max_row)
    print("=========清空过期数据========")
    columnI = 2
    sheet.cell(row=1, column=columnI).value = "类型"
    if file1 == "unitOrigin.ini":
        sheet.cell(row=2, column=columnI).value = "unit"
    elif file1 == "abilityOrigin.ini":
        sheet.cell(row=2, column=columnI).value = "ability"
    elif file1 == "itemOrigin.ini":
        sheet.cell(row=2, column=columnI).value = "item"
    elif file1 == "upgradeOrigin.ini":
        sheet.cell(row=2, column=columnI).value = "upgrade"
        

    # 列

    columnI = columnI + 1
    sheet.cell(row=1, column=columnI).value = "物遍ID"
    sheet.cell(row=2, column=columnI).value = "ID"
    columnI = columnI + 1
    sheet.cell(row=1, column=columnI).value = "父对象"
    sheet.cell(row=2, column=columnI).value = "parent"

    rowI = 2
    offsetColunm = 5
    UbertipStr = ""
    UbertipStrBoolean = False
    UbertipcolumnI = 0

    longDataStr = ""
    longDataColumnI = 0

    titleTemp = ""
    titleCodeDes = []
    titleCode = []
    with open(file1, 'rb') as f:
        list = f.readlines()
        titleCodeBoolean = False
        for line in list:
            line = line.strip()
            str = line.decode('utf-8')

            # 写入ID
            if "[" in str and "]" in str and len(str) == 6:
                # 清空UbertipStr
                if UbertipStrBoolean:
                    UbertipStrBoolean = False
                    sheet.cell(
                        row=rowI, column=UbertipcolumnI).value = UbertipStr
                    UbertipStr = ""
                # if not longDataColumnI == 0:
                #     sheet.cell(row=rowI, column=longDataColumnI ).value = longDataStr
                #     longDataColumnI = 0
                #     longDataStr = ""

                str = str[1:-1]
                # 第几行
                rowI += 1
                sheet.cell(row=rowI, column=3).value = str
            # 父对象
            elif "_parent" in str:
                startI = str.find("\"")
                str2 = str[startI:]
                sheet.cell(row=rowI, column=4).value = str2
            # 静态UI东西 技能
            elif "W2LObject" in str:
                startI = 0
                endI = str.find("=") - 1
                # 标题
                str2 = str[startI:endI]
                 # 值
                str3 = str[endI+3:]
                if not str2 in titleCode:
                    columnI += 1
                    titleCode.append(str2)
                    sheet.cell(row=1, column=columnI).value = "静态对象"
                    sheet.cell(row=2, column=columnI).value = str2
                    sheet.cell(row=rowI, column= columnI ).value = str3
                else:
                    i = titleCode.index(str2)
                    sheet.cell(row=rowI, column=i+offsetColunm ).value = str3
            elif str[0:2] == "--":
                # 清空UbertipStr
                if UbertipStrBoolean:
                    UbertipStrBoolean = False
                    sheet.cell(row=rowI, column=UbertipcolumnI).value = UbertipStr
                    UbertipStr = ""

                str2 = str[3:]
                titleCodeBoolean = True
                if not str2 in titleCodeDes:
                    titleTemp = str2
                    titleCodeDes.append(str2)
            
            elif titleCodeBoolean:
                titleCodeBoolean = False
                startI = 0
                endI = str.find("=") - 1
                # 标题
                str2 = str[startI:endI]
                # 值
                str3 = str[endI+3:]

                if not str2 in titleCode:
                    titleCode.append(str2)
                    columnI += 1
                    # 技能里会出现同个字段 但是注释不同 如DataA DataB DataC
                    if str2 == "DataA":
                        titleTemp = "参数A"
                    elif str2 == "DataB":
                        titleTemp = "参数B"
                    elif str2 == "DataC":
                        titleTemp = "参数C"
                    elif str2 == "DataD":
                        titleTemp = "参数D"
                    elif str2 == "DataE":
                        titleTemp = "参数E"
                    sheet.cell(row=1, column=columnI).value = titleTemp
                    sheet.cell(row=2, column=columnI).value = str2
                    # 特殊处理
                    if str2 == "Ubertip":
                        UbertipStr += str3
                        UbertipStrBoolean = True
                        UbertipcolumnI = columnI
                    elif str3[0] == "{" and str3[-1] != "}":
                        longDataStr += str3
                        longDataColumnI = columnI
                    else:
                        sheet.cell(row=rowI, column=columnI).value = str3
                else:
                    # 新增
                    i = titleCode.index(str2)
                    # 特殊处理
                    if str2 == "Ubertip":
                        UbertipStr += str3
                        UbertipStrBoolean = True
                        UbertipcolumnI = i+offsetColunm
                    elif str3[0] == "{" and str3[-1] != "}":
                        longDataStr += str3
                        longDataColumnI = i+offsetColunm
                    else:
                        sheet.cell(row=rowI, column=i +
                                   offsetColunm).value = str3
            elif UbertipStrBoolean:
                UbertipStr += str
            elif longDataColumnI != 0:
                longDataStr += str
            elif len(str)>0 and str[-1] == "}":
                if not longDataColumnI == 0:
                    sheet.cell(
                        row=rowI, column=longDataColumnI).value = longDataStr
                    longDataColumnI = 0
                    longDataStr = ""
                else:
                    assert False


wb = openpyxl.load_workbook('demo2.xlsm', keep_vba=True)

file1 = r"unitOrigin.ini"
if(os.path.exists(file1)):
    print("=========开始导出单位数据========")
    writeDoo(file1,wb)
    print("=========单位数据导出完毕========")

file1 = r"abilityOrigin.ini"
if(os.path.exists(file1)):
    print("=========开始导出技能数据========")
    writeDoo(file1, wb)
    print("=========技能数据导出完毕========")

file1 = r"itemOrigin.ini"
if(os.path.exists(file1)):
    print("=========开始导出物品数据========")
    writeDoo(file1, wb)
    print("=========物品数据导出完毕========")

file1 = r"upgradeOrigin.ini"
if(os.path.exists(file1)):
    print("=========开始导出科技数据========")
    writeDoo(file1, wb)
    print("=========科技数据导出完毕========")

print("=========正在保存========")
wb.save('demo2.xlsm')  # 保存文件，注意以xlsx为文件扩展名
print("=========保存完毕！========")

# writeXlsx()
